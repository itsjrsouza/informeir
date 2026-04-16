#!/usr/bin/env python3
"""
parseExcel.py
Lê o Excel de importação e retorna JSON com os dados de cada sócio.
Uso: python3 parseExcel.py <arquivo.xlsx>
"""

import sys
import json
import re
import os
from openpyxl import load_workbook

CAMPOS_NUMERICOS = {
    "tributaveis", "inss", "outrasDeducoes", "irrf",
    "lucrosDividendos", "outrosIsentos",
    "planoBeneficiario", "planoFontePagadora",
    "trezeRendimentos", "trezeInss", "trezeIrrf",
}

# Mapeamento: nome da coluna no Excel → chave interna
COL_MAP = {
    "ANO_CALENDARIO":      ("anoCalendario",      "meta"),
    "CNPJ_EMPRESA":        ("cnpj",               "fontePagadora"),
    "RAZAO_SOCIAL":        ("razaoSocial",         "fontePagadora"),
    "NOME_SOCIO":          ("nome",                "beneficiario"),
    "CPF_SOCIO":           ("cpf",                 "beneficiario"),
    "TRIBUTAVEIS":         ("tributaveis",         "rendimentos"),
    "INSS":                ("inss",                "rendimentos"),
    "OUTRAS_DEDUCOES":     ("outrasDeducoes",      "rendimentos"),
    "IRRF":                ("irrf",                "rendimentos"),
    "LUCROS_DIVIDENDOS":   ("lucrosDividendos",    "rendimentos"),
    "OUTROS_ISENTOS":      ("outrosIsentos",       "rendimentos"),
    "PLANO_BENEFICIARIO":  ("planoBeneficiario",   "rendimentos"),
    "PLANO_FONTE":         ("planoFontePagadora",  "rendimentos"),
    "TREZE_RENDIMENTOS":   ("trezeRendimentos",    "rendimentos"),
    "TREZE_INSS":          ("trezeInss",           "rendimentos"),
    "TREZE_IRRF":          ("trezeIrrf",           "rendimentos"),
}

CAMPOS_OBRIGATORIOS = ["CNPJ_EMPRESA", "RAZAO_SOCIAL", "NOME_SOCIO", "CPF_SOCIO", "ANO_CALENDARIO"]

def limpar_numero(val):
    if val is None or val == "":
        return "0,00"
    s = str(val).strip()
    # Remove R$, espaços
    s = re.sub(r"[R$\s]", "", s)
    # Se tem vírgula e ponto: formato brasileiro 1.234,56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        f = float(s)
        # Devolve no formato "1234,56" para o gerador JS
        return f"{f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"

def limpar_cnpj(val):
    return re.sub(r"\D", "", str(val or ""))

def limpar_cpf(val):
    return re.sub(r"\D", "", str(val or ""))

def verificar_arquivo(caminho):
    """Verificações iniciais do arquivo"""
    if not os.path.exists(caminho):
        return {"erro": f"Arquivo não encontrado: {caminho}"}
    
    tamanho = os.path.getsize(caminho)
    if tamanho == 0:
        return {"erro": f"Arquivo vazio (0 bytes)"}
    
    # Verificar assinatura do arquivo
    with open(caminho, 'rb') as f:
        header = f.read(4)
        if header[:2] != b'PK':
            # Mostrar preview do conteúdo para diagnóstico
            f.seek(0)
            preview = f.read(200).decode('utf-8', errors='ignore')
            return {
                "erro": f"Arquivo não é um .xlsx válido (header: {header.hex()})",
                "tamanho": tamanho,
                "preview": preview[:100]
            }
    
    return {"ok": True, "tamanho": tamanho}

def parse(caminho):
    # Verificações iniciais
    verificacao = verificar_arquivo(caminho)
    if "erro" in verificacao:
        return verificacao
    
    try:
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
    except Exception as e:
        return {
            "erro": f"Erro ao abrir Excel: {str(e)}",
            "tipo": type(e).__name__
        }

    # Detecta linha de cabeçalho (primeira linha que tem "ANO_CALENDARIO")
    header_row = None
    header_map = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_idx, cell in enumerate(row):
            if str(cell or "").strip() == "ANO_CALENDARIO":
                header_row = row_idx
                for ci, h in enumerate(row):
                    if h:
                        header_map[str(h).strip()] = ci
                break
        if header_row:
            break

    if header_row is None:
        return {"erro": "Coluna ANO_CALENDARIO não encontrada. Verifique se está usando o modelo correto."}

    socios = []
    erros = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Pula linhas completamente vazias
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # Pula linha de label amigável (segunda linha do cabeçalho no modelo)
        primeira_cel = str(row[0] or "").strip()
        if primeira_cel in ("", "Ano-Calendário"):
            continue

        socio = {"fontePagadora": {}, "beneficiario": {}, "rendimentos": {}}
        erros_linha = []

        for col_nome, (chave, secao) in COL_MAP.items():
            if col_nome not in header_map:
                continue
            val = row[header_map[col_nome]]

            if secao == "meta":
                try:
                    socio["anoCalendario"] = int(val) if val else None
                except:
                    socio["anoCalendario"] = None
            elif secao == "fontePagadora":
                if col_nome == "CNPJ_EMPRESA":
                    socio["fontePagadora"][chave] = limpar_cnpj(val)
                else:
                    socio["fontePagadora"][chave] = str(val or "").strip().upper()
            elif secao == "beneficiario":
                if col_nome == "CPF_SOCIO":
                    socio["beneficiario"][chave] = limpar_cpf(val)
                else:
                    socio["beneficiario"][chave] = str(val or "").strip().upper()
            elif secao == "rendimentos":
                socio["rendimentos"][chave] = limpar_numero(val)

        # Validações obrigatórias
        if not socio["fontePagadora"].get("cnpj"):
            erros_linha.append("CNPJ_EMPRESA vazio")
        if not socio["fontePagadora"].get("razaoSocial"):
            erros_linha.append("RAZAO_SOCIAL vazia")
        if not socio["beneficiario"].get("nome"):
            erros_linha.append("NOME_SOCIO vazio")
        if not socio["beneficiario"].get("cpf"):
            erros_linha.append("CPF_SOCIO vazio")
        if not socio.get("anoCalendario"):
            erros_linha.append("ANO_CALENDARIO inválido")

        if erros_linha:
            erros.append({"linha": row_idx, "erros": erros_linha, "nome": socio["beneficiario"].get("nome", "?")})
            continue

        # Responsável fixo
        socio["responsavel"] = {
            "nome": "NELSON FERNANDES DE SOUZA JUNIOR",
            "data": "28.02.2026"
        }

        socios.append(socio)

    return {
        "total": len(socios),
        "registros": socios,
        "validos": len(socios),
        "invalidos": len(erros),
        "erros": erros,
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"erro": "Informe o caminho do arquivo Excel"}))
        sys.exit(1)
    
    result = parse(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))