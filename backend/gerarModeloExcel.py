#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

COLUNAS = [
    ("EXERCICIO",                   "Exercício (ano de entrega)",            12),
    ("ANO_CALENDARIO",              "Ano-Calendário",                        12),
    ("CNPJ_EMPRESA",                "CNPJ da Empresa",                       20),
    ("RAZAO_SOCIAL",                "Razão Social / Nome Empresarial",       35),
    ("NOME_BENEFICIARIO",           "Nome Completo do Beneficiário",         35),
    ("CPF_BENEFICIARIO",            "CPF do Beneficiário",                   16),
    ("TRIBUTAVEIS",                 "01-Total Rendimentos Tributáveis",      22),
    ("INSS",                        "02-Contribuição Previdenciária",        20),
    ("PREV_COMPLEMENTAR",           "03-Previdência Complementar/FAPI",      20),
    ("PENSAO_ALIMENTICIA",          "04-Pensão Alimentícia",                 20),
    ("IRRF",                        "05-IRRF",                               16),
    ("PARCELA_ISENTA65",            "01-Parcela isenta aposentadoria, pensão e reforma (65 anos ou mais)",                 20),
    ("PARCELA_13",                  "02-Parcela isenta do 13º de aposentadoria, pensão e reforma (65 anos ou mais)",                 20),
    ("DIARIAS_AJUDAS",              "02-Diárias e Ajudas de Custo",          20),
    ("MOLESTIA_GRAVE",              "03-Pensão/aposentadoria/reforma por moléstia grave ou acidente em serviço",            20),
    ("LUCROS_DIVIDENDOS",           "04-Lucros e Dividendos",                22),
    ("LUCROS_SIMPLES_NACIONAL",     "05-Valores pagos a titular ou sócio de ME/EPP (exceto pró-labore, aluguéis e serviços)",           22),
    ("INDENIZACOES",                "06-Indenizações/PDV",                   20),
    ("OUTROS_ISENTOS",              "07-Outros Rendimentos Isentos",         22),
    ("TREZE_RENDIMENTOS",           "01-13º Salário",                        20),
    ("TREZE_IRRF",                  "02-IRRF sobre 13º",                     20),
    ("OUTROS_TRIB_EXCL",            "03-Outros Trib. Exclusiva",             20),
    ("RRA_NUM_PROCESSO",            "RRA - Número do Processo",              25),
    ("RRA_MESES",                   "RRA - Quantidade de Meses",             15),
    ("NATUREZA_RENDIMENTO",         "Natureza do Rendimento",                25),
    ("NATUREZA_RENDIMENTO_RRA",     "Natureza do Rendimento (RRA)",          25),
    ("RRA_TRIBUTAVEIS",             "RRA - Total Rendimentos Tributáveis",   22),
    ("RRA_DESP_JUDICIAIS",          "RRA - Despesas Judiciais",              20),
    ("RRA_INSS",                    "RRA - Contribuição Previdenciária",     20),
    ("RRA_PENSAO",                  "RRA - Pensão Alimentícia",              20),
    ("RRA_IRRF",                    "RRA - IRRF",                            16),
    ("RRA_ISENTOS",                 "RRA - Rendimentos Isentos",             22),
    ("INFO_COMPLEMENTARES",         "Informações Complementares",            40),
    ("RESP_NOME",                   "Responsável - Nome",                    30),
    ("RESP_DATA",                   "Responsável - Data",                    15),
    ("RESP_ASSINATURA",             "Responsável - Assinatura",              30),
]

EXEMPLOS = [
    [2026, 2025, "12.345.678/0001-95", "EMPRESA ALPHA LTDA", "JOÃO DA SILVA SOUZA", "123.456.789-01",
     84000.00, 7786.00, 0.00, 0.00, 4200.00,
     0.00, 0.00, 0.00, 0.00, 0.00, 120000.00, 0.00, 0.00,
     7000.00, 0.00, 0.00,
     "", "", "Assalariado", "Rendimentos acumulados", 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
     "Nada a declarar", "Contador XYZ", "28/02/2026", "Isento conforme IN RFB 1.215/2011"],
]

AZUL_ESCURO, AZUL_HEADER, AZUL_CLARO, VERDE_EX, BRANCO, CINZA = "1F3864", "2563A8", "D9E1F2", "E2EFDA", "FFFFFF", "F5F5F5"
def borda(cor="CCCCCC"): s=Side(style="thin", color=cor); return Border(left=s, right=s, top=s, bottom=s)

def gerar(output_path):
    wb = Workbook()
    ws = wb.active; ws.title = "MODELO"
    ws.merge_cells("A1:AO1"); ws["A1"].value = "IMPORTAÇÃO EM LOTE — INFORME DE RENDIMENTOS"
    ws["A1"].font = Font(bold=True, size=14, color=BRANCO); ws["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:AO2"); ws["A2"].value = "Preencha uma linha por beneficiário. Não altere os nomes das colunas."
    ws["A2"].font = Font(italic=True, size=10, color="444444"); ws["A2"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for col_idx, (campo, label, largura) in enumerate(COLUNAS, start=1):
        col_letra = get_column_letter(col_idx)
        c = ws.cell(row=3, column=col_idx, value=campo); c.font = Font(bold=True, size=10, color=BRANCO)
        c.fill = PatternFill("solid", fgColor=AZUL_HEADER); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda("1F3864")
        ws.column_dimensions[col_letra].width = largura
    ws.row_dimensions[3].height = 22

    for row_idx, exemplo in enumerate(EXEMPLOS, start=4):
        for col_idx, valor in enumerate(exemplo, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=valor); c.font = Font(size=10, color="1a5276")
            c.fill = PatternFill("solid", fgColor=VERDE_EX); c.border = borda()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if 7 <= col_idx <= 32: c.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 18

    for row_idx in range(5, 105):
        for col_idx in range(1, len(COLUNAS)+1):
            c = ws.cell(row=row_idx, column=col_idx); c.border = borda()
            c.fill = PatternFill("solid", fgColor=BRANCO if row_idx % 2 == 0 else CINZA)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if 7 <= col_idx <= 32: c.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A4"
    wb.save(output_path)
    print(f"✅ Modelo Excel gerado: {output_path}")

if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "modelo.xlsx"
    gerar(out)